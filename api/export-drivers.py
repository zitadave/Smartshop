import json
import requests
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Supabase Configurations
SUPABASE_URL = "https://auaendcgszofgvdfdajt.supabase.co"
SUPABASE_KEY = "" # Will be resolved from process environment

def handler(request):
    # Fetch drivers from Supabase directly
    headers = {
        "apikey": SUPABASE_KEY or "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9...", # fallback or env
        "Authorization": f"Bearer {SUPABASE_KEY}"
    }
    
    # Try to load credentials from Vercel env or python environment
    env_key = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("SUPABASE_KEY")
    if env_key:
        headers = {
            "apikey": env_key,
            "Authorization": f"Bearer {env_key}"
        }

    # Fetch data
    try:
        r = requests.get(f"{SUPABASE_URL}/rest/v1/delivery_personnel?order=joined_at.desc", headers=headers)
        drivers = r.json() if r.status_code == 200 else []
    except Exception as e:
        drivers = []

    # Create openpyxl Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Personnel"

    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True

    # Styling Palettes
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid") # Emerald Green
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, bold=False, color="000000")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # Column Headers
    headers_list = [
        "Driver ID", "Full Name (Latin)", "Full Name (Amharic)", "Phone", "Email", 
        "Fayda ID", "Vehicle Type", "License Plate", "Status", "Rating", 
        "Deliveries", "Earnings (Br)", "Score", "Tier", "Online", "Joined At"
    ]

    ws.append(headers_list)

    # Style Header Row
    ws.row_dimensions[1].height = 28
    for col_num, header_title in enumerate(headers_list, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Populate Data
    for row_idx, d in enumerate(drivers, 2):
        row_data = [
            d.get("id"),
            d.get("full_name_latin", ""),
            d.get("full_name_amharic", ""),
            d.get("phone", ""),
            d.get("email", ""),
            d.get("fayda_id", ""),
            d.get("vehicle_type", "motorcycle"),
            d.get("license_plate", ""),
            d.get("status", "pending_review"),
            float(d.get("rating") or 0.0),
            int(d.get("total_deliveries") or 0),
            int(d.get("total_earnings") or 0),
            int(d.get("driver_score") or 0),
            d.get("driver_tier", "bronze"),
            "Yes" if d.get("is_online") else "No",
            d.get("joined_at", "")[:10] if d.get("joined_at") else ""
        ]
        ws.append(row_data)
        
        # Row dimensions and styles
        ws.row_dimensions[row_idx].height = 20
        use_zebra = (row_idx % 2 == 1)
        
        for col_idx in range(1, len(headers_list) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if use_zebra:
                cell.fill = zebra_fill
            
            # Alignments
            if col_idx in [1, 4, 6, 8, 9, 14, 15, 16]:
                cell.alignment = align_center
            elif col_idx in [10, 11, 12, 13]:
                cell.alignment = align_right
            else:
                cell.alignment = align_left

    # Auto-fit Column Widths cleanly
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # Save to binary stream
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    xlsx_bytes = file_stream.getvalue()

    # Dispatch compiled report to admin Telegram chat
    bot_token = os.environ.get("TELEGRAM_ADMIN_BOT_TOKEN") or "8951025148:AAG456KIIBnyLBQqbkeDLajcT_TaPSYCIYc"
    chat_id = os.environ.get("TELEGRAM_ADMIN_CHAT_ID") or "336997351"
    
    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
    files = {
        'document': ('drivers_report.xlsx', xlsx_bytes, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    }
    data_payload = {
        'chat_id': chat_id,
        'caption': "📊 *Smartshop Express Delivery Fleet Report*\n\nHere is the requested Excel spreadsheet of all driver applications, statuses, earnings, and ratings."
    }
    
    success = False
    try:
        res_tg = requests.post(url, data=data_payload, files=files)
        success = (res_tg.status_code == 200)
    except Exception as e:
        print("Telegram sending error:", e)

    # Return response as JSON
    return {
        "statusCode": 200,
        "headers": {
            "Content-Type": "application/json",
            "Cache-Control": "no-cache"
        },
        "body": json.dumps({"success": success, "message": "Fleet report sent directly to your Telegram admin chat!"})
    }
